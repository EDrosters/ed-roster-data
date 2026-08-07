"""
refresh_roster.py

Runs in GitHub Actions on a schedule. Steps:
  1. Uses the stored refresh token to get a fresh Microsoft Graph access token
     (and stores the NEW refresh token back into GitHub Secrets, since Microsoft
     rotates them on every use).
  2. Downloads the roster Excel file from OneDrive (found by filename search,
     so it keeps working even if the file moves folders).
  3. Runs the same parsing logic we built and tested locally to produce
     roster_data.json (the Today + Tomorrow view).
  4. Commits roster_data.json back into this repo, where GitHub Pages serves it.
"""

import os, re, sys, json, base64
from io import BytesIO
from datetime import datetime, date, timedelta
from zoneinfo import ZoneInfo

import requests
import openpyxl

# ============================================================
# CONFIG - filled in from GitHub Secrets / environment
# ============================================================
CLIENT_ID = os.environ["MS_CLIENT_ID"]
REFRESH_TOKEN = os.environ["MS_REFRESH_TOKEN"]
GH_PAT = os.environ.get("GH_PAT")                     # needed to rotate the refresh token secret
GH_REPOSITORY = os.environ.get("GITHUB_REPOSITORY")   # auto-set by Actions, e.g. "user/ed-roster-data"

ROSTER_FILENAME_SEARCH = os.environ.get("ROSTER_FILENAME", "Master Roster")

MELBOURNE_TZ = ZoneInfo("Australia/Melbourne")

def local_today():
    """Melbourne's current calendar date - NOT the server's UTC date.
    GitHub Actions runners run on UTC, and Melbourne is 10-11 hours ahead,
    so using the server's own date.today() would show yesterday's roster
    for several hours every morning."""
    return datetime.now(MELBOURNE_TZ).date()

TODAY = local_today().isoformat()
TOMORROW = (local_today() + timedelta(days=1)).isoformat()

# ============================================================
# STEP 1 - Refresh the access token (and rotate the stored secret)
# ============================================================

def get_access_token():
    resp = requests.post(
        "https://login.microsoftonline.com/consumers/oauth2/v2.0/token",
        data={
            "client_id": CLIENT_ID,
            "grant_type": "refresh_token",
            "refresh_token": REFRESH_TOKEN,
            "scope": "Files.Read offline_access",
        },
    )
    resp.raise_for_status()
    payload = resp.json()
    new_refresh_token = payload.get("refresh_token")
    if new_refresh_token and new_refresh_token != REFRESH_TOKEN:
        update_github_secret("MS_REFRESH_TOKEN", new_refresh_token)
    return payload["access_token"]


def update_github_secret(name, value):
    """Encrypts and stores a new value for a repo secret via the GitHub API."""
    if not GH_PAT or not GH_REPOSITORY:
        print("::warning::No GH_PAT set - cannot rotate refresh token. "
              "The stored MS_REFRESH_TOKEN may go stale.")
        return
    try:
        from nacl import encoding, public
    except ImportError:
        print("::warning::PyNaCl not installed - cannot rotate refresh token.")
        return

    headers = {"Authorization": f"Bearer {GH_PAT}", "Accept": "application/vnd.github+json"}
    key_resp = requests.get(
        f"https://api.github.com/repos/{GH_REPOSITORY}/actions/secrets/public-key",
        headers=headers,
    )
    key_resp.raise_for_status()
    key_data = key_resp.json()

    public_key = public.PublicKey(key_data["key"].encode("utf-8"), encoding.Base64Encoder())
    sealed_box = public.SealedBox(public_key)
    encrypted = sealed_box.encrypt(value.encode("utf-8"))
    encrypted_b64 = base64.b64encode(encrypted).decode("utf-8")

    put_resp = requests.put(
        f"https://api.github.com/repos/{GH_REPOSITORY}/actions/secrets/{name}",
        headers=headers,
        json={"encrypted_value": encrypted_b64, "key_id": key_data["key_id"]},
    )
    put_resp.raise_for_status()
    print(f"Rotated secret {name} successfully.")


# ============================================================
# STEP 2 - Download the roster file from OneDrive
# ============================================================

def download_roster(access_token):
    headers = {"Authorization": f"Bearer {access_token}"}
    search_resp = requests.get(
        f"https://graph.microsoft.com/v1.0/me/drive/root/search(q='{ROSTER_FILENAME_SEARCH}')",
        headers=headers,
    )
    search_resp.raise_for_status()
    items = search_resp.json().get("value", [])
    xlsx_items = [i for i in items if i["name"].lower().endswith(".xlsx")]
    if not xlsx_items:
        raise RuntimeError(f"No .xlsx file found matching '{ROSTER_FILENAME_SEARCH}'")
    item = xlsx_items[0]
    print(f"Found roster file: {item['name']} (id={item['id']})")

    content_resp = requests.get(
        f"https://graph.microsoft.com/v1.0/me/drive/items/{item['id']}/content",
        headers=headers,
    )
    content_resp.raise_for_status()
    return BytesIO(content_resp.content)


# ============================================================
# STEP 3 - Parsing logic (consultant / registrar / JMO / NP / AMP tabs)
# ============================================================

DATE_TAB_RE = re.compile(r"^[0-9.\-]+$")
THEME_COLORS = ['FFFFFF', '000000', 'EEECE1', '1F497D', '4F81BD', 'C0504D', '9BBB59', '8064A2', '4BACC6', 'F79646']

def apply_tint(hexcol, tint):
    r = int(hexcol[0:2], 16); g = int(hexcol[2:4], 16); b = int(hexcol[4:6], 16)
    if tint < 0:
        r *= (1 + tint); g *= (1 + tint); b *= (1 + tint)
    elif tint > 0:
        r = r + (255 - r) * tint
        g = g + (255 - g) * tint
        b = b + (255 - b) * tint
    return '%02X%02X%02X' % (round(max(0, min(255, r))), round(max(0, min(255, g))), round(max(0, min(255, b))))

def soften(hexcol, amount=0.55):
    r = int(hexcol[0:2], 16); g = int(hexcol[2:4], 16); b = int(hexcol[4:6], 16)
    r = r + (255 - r) * amount
    g = g + (255 - g) * amount
    b = b + (255 - b) * amount
    return '%02X%02X%02X' % (round(r), round(g), round(b))

def resolve_fill(cell):
    fill = cell.fill
    if not fill or fill.patternType != "solid":
        return None
    fg = fill.fgColor
    hexcol = None
    if fg.type == "rgb" and fg.rgb and isinstance(fg.rgb, str) and len(fg.rgb) == 8:
        if fg.rgb[:2] != "00":
            hexcol = fg.rgb[2:]
    elif fg.type == "theme":
        try:
            base = THEME_COLORS[fg.theme]
            hexcol = apply_tint(base, fg.tint or 0)
        except Exception:
            hexcol = None
    if not hexcol:
        return None
    r = int(hexcol[0:2], 16) / 255; g = int(hexcol[2:4], 16) / 255; b = int(hexcol[4:6], 16) / 255
    lum = 0.2126 * r + 0.7152 * g + 0.0722 * b
    if lum > 0.96:
        return None
    return soften(hexcol)
METRIC_LABELS = {
    "facem total", "registrars", "jmos rostered", "jmos minimum",
    "director on call", "funded", "reg sick call",
}
METRIC_KEY_MAP = {
    "facem total": "facem", "registrars": "reg",
    "jmos rostered": "jmo", "director on call": "oncall",
}
REG_TABS = ["Reg T1 2025", "Reg T2 2025", "Reg T3 2025", "Reg T4 2025",
            "Reg T1 2026", "Reg T2 2026", "Reg T3 2026"]
TERM_LABELS = {
    "Reg T1 2025": "Term 1, 2025", "Reg T2 2025": "Term 2, 2025",
    "Reg T3 2025": "Term 3, 2025", "Reg T4 2025": "Term 4, 2025",
    "Reg T1 2026": "Term 1, 2026", "Reg T2 2026": "Term 2, 2026",
    "Reg T3 2026": "Term 3, 2026",
}

def norm(s):
    return re.sub(r"\s+", " ", str(s)).strip().lower()

def fmt_code(v):
    if v is None:
        return ""
    if isinstance(v, str) and v.strip() in ("", "\xa0"):
        return ""
    return str(v).strip()

def find_date_anchor(ws, max_row=6, max_col=8):
    for r in range(1, max_row):
        for c in range(2, max_col):
            v = ws.cell(row=r, column=c).value
            if isinstance(v, datetime):
                return r, c
    return None, None

def find_last_date_col(ws, date_row, start_col, limit=270):
    last = start_col
    for c in range(start_col, limit):
        v = ws.cell(row=date_row, column=c).value
        if isinstance(v, datetime):
            last = c
    return last


def build_weekday_base(wb):
    weekday_base = {}
    data_ws = wb["Data"]
    for r in range(2, 9):
        wd = data_ws.cell(row=r, column=1).value
        day_base = data_ws.cell(row=r, column=2).value
        eve_base = data_ws.cell(row=r, column=3).value
        if wd:
            weekday_base[str(wd).strip()] = (day_base or 0) + (eve_base or 0)
    return weekday_base

def to_num(v):
    try:
        return float(v)
    except (TypeError, ValueError):
        return None

def facem_status(actual_total, weekday, weekday_base):
    base = weekday_base.get(weekday)
    if actual_total is None or base is None:
        return ""
    diff = actual_total - base
    if diff > 0:
        return "b"
    if diff == 0:
        return "g"
    if diff == -1:
        return "y"
    return "r"


def extract_consultant_tab(ws, name, weekday_base):
    date_row, start_col = find_date_anchor(ws, 6, 7)
    if date_row is None:
        return None
    weekday_row = date_row + 1
    last_col = find_last_date_col(ws, date_row, start_col, 46)

    dates, weekdays = [], []
    for c in range(start_col, last_col + 1):
        dv = ws.cell(row=date_row, column=c).value
        wv = ws.cell(row=weekday_row, column=c).value
        dates.append(dv.strftime("%Y-%m-%d") if isinstance(dv, datetime) else None)
        weekdays.append(str(wv).strip() if wv else "")

    r = weekday_row + 1
    metrics = {}
    while True:
        b = ws.cell(row=r, column=2).value
        if b is not None and norm(b) in METRIC_LABELS:
            values = [fmt_code(ws.cell(row=r, column=c).value) for c in range(start_col, last_col + 1)]
            metrics[str(b).strip()] = values
            r += 1
            continue
        break

    consultants = []
    current_group = None
    blank_streak = 0
    dse_rows = {}
    while r < ws.max_row:
        a = ws.cell(row=r, column=1).value
        b = ws.cell(row=r, column=2).value
        if a not in (None, ""):
            current_group = str(a).strip()
        if b in (None, ""):
            blank_streak += 1
            if blank_streak >= 5:
                break
            r += 1
            continue
        if norm(b) in ("d", "s", "e"):
            dse_rows = {}
            for offset in range(3):
                lbl = ws.cell(row=r + offset, column=2).value
                if lbl and norm(lbl) in ("d", "s", "e"):
                    vals = [to_num(ws.cell(row=r + offset, column=c).value) for c in range(start_col, last_col + 1)]
                    dse_rows[norm(lbl)] = vals
            break
        blank_streak = 0
        codes = []
        fills = []
        for c in range(start_col, last_col + 1):
            cell = ws.cell(row=r, column=c)
            codes.append(fmt_code(cell.value))
            fills.append(resolve_fill(cell) or "")
        consultants.append({"n": str(b).strip(), "g": current_group or "", "v": "|".join(codes), "f": "|".join(fills)})
        r += 1

    metrics_out = {}
    for label, values in metrics.items():
        key = METRIC_KEY_MAP.get(norm(label))
        if key:
            metrics_out[key] = "|".join(values)

    facem_status_list = []
    d_vals = dse_rows.get("d", [])
    s_vals = dse_rows.get("s", [])
    e_vals = dse_rows.get("e", [])
    for i, wd in enumerate(weekdays):
        d = d_vals[i] if i < len(d_vals) else None
        s = s_vals[i] if i < len(s_vals) else None
        e = e_vals[i] if i < len(e_vals) else None
        parts = [x for x in (d, s, e) if x is not None]
        total = sum(parts) if parts else None
        facem_status_list.append(facem_status(total, wd, weekday_base) if total is not None else "")
    metrics_out["facem_status"] = "|".join(facem_status_list)

    start_date = next((d for d in dates if d), None)
    end_date = next((d for d in reversed(dates) if d), None)
    return {"t": name, "s": start_date, "e": end_date,
            "d": "|".join(d or "" for d in dates), "w": "|".join(weekdays),
            "m": metrics_out, "c": consultants}


def extract_all_consultants(wb):
    weekday_base = build_weekday_base(wb)
    periods = []
    for name in wb.sheetnames:
        if DATE_TAB_RE.match(name.strip()):
            result = extract_consultant_tab(wb[name], name, weekday_base)
            if result:
                periods.append(result)
    periods.sort(key=lambda p: p["s"] or "")
    return periods


def extract_registrar_term(ws, tab_name):
    date_row, start_col = find_date_anchor(ws, 6, 8)
    if date_row is None:
        return None
    last_col = find_last_date_col(ws, date_row, start_col, 260)
    dates, weekdays = [], []
    for c in range(start_col, last_col + 1):
        dv = ws.cell(row=date_row, column=c).value
        dates.append(dv.strftime("%Y-%m-%d") if isinstance(dv, datetime) else None)
        weekdays.append(dv.strftime("%a") if isinstance(dv, datetime) else "")

    r = date_row + 1
    registrars = []
    while r <= ws.max_row:
        a = ws.cell(row=r, column=1).value
        if a is not None and str(a).strip().upper() == "TOTAL":
            break
        if a in (None, ""):
            r += 1
            continue
        name = str(a).strip()
        eft = ws.cell(row=r, column=2).value
        codes = [fmt_code(ws.cell(row=r, column=c).value) for c in range(start_col, last_col + 1)]
        registrars.append({"n": name, "eft": eft, "v": "|".join(codes)})
        r += 1

    registrars.sort(key=lambda x: (-(x["eft"] or 0), x["n"]))

    start_date = next((d for d in dates if d), None)
    end_date = next((d for d in reversed(dates) if d), None)
    return {"t": TERM_LABELS.get(tab_name, tab_name), "s": start_date, "e": end_date,
            "d": "|".join(d or "" for d in dates), "w": "|".join(weekdays), "c": registrars}


def extract_all_registrars(wb):
    periods = []
    for tab in REG_TABS:
        if tab in wb.sheetnames:
            result = extract_registrar_term(wb[tab], tab)
            if result:
                periods.append(result)
    periods.sort(key=lambda p: p["s"] or "")
    return periods


def build_date_col_map(ws, date_row, col_start=2, col_end=200):
    """Some sheets repeat the same calendar date in more than one column
    (a duplicate/leftover week block). The first occurrence consistently
    holds the real data and later duplicates are blank, so keep the first
    one found rather than letting a later duplicate silently overwrite it."""
    m = {}
    for c in range(col_start, col_end):
        v = ws.cell(row=date_row, column=c).value
        if isinstance(v, datetime):
            key = v.strftime("%Y-%m-%d")
            if key not in m:
                m[key] = c
    return m

def is_legend_row(name):
    """Some sheets embed a code legend/reference table below the real staff
    list (e.g. 'Role/Code', 'DO1 (SSU 1-12)', 'D = 0800 - 1630'). These aren't
    real people - stop scanning as soon as one is hit."""
    low = name.strip().lower()
    if low == "role/code":
        return True
    if " = " in name:
        return True
    return False

def extract_name_rows(ws, date_col_map, name_col, row_start, row_end):
    people = {}
    for r in range(row_start, row_end):
        name = fmt_code(ws.cell(row=r, column=name_col).value)
        if not name:
            continue
        if is_legend_row(name):
            break
        codes = {}
        for date_str, col in date_col_map.items():
            code = fmt_code(ws.cell(row=r, column=col).value)
            if code and code != "0":
                codes[date_str] = code
        people[name] = codes
    return people

def extract_jmo_np(wb):
    ws_jmo = wb["JMO"]
    jmo_dates = build_date_col_map(ws_jmo, 2, 2, 260)
    jmo_people = extract_name_rows(ws_jmo, jmo_dates, 1, 5, 545)

    ws_np = wb["NP"]
    np_dates = build_date_col_map(ws_np, 1, 2, 70)
    np_people = extract_name_rows(ws_np, np_dates, 1, 3, 40)

    ws_amp = wb["AMP2"]
    amp_dates = build_date_col_map(ws_amp, 2, 2, 65)
    amp_people = extract_name_rows(ws_amp, amp_dates, 1, 3, 22)

    return jmo_people, np_people, amp_people


# ============================================================
# STEP 4 - Shift/zone/role template (hard-coded; matches the
# original "Weekly Allocations" tab layout as of the last time
# it existed - update here if position slots ever change)
# ============================================================

TEMPLATE = [{'shift': 'Day', 'zone': None, 'role': 'FACEM on call', 'time': '0430-0800'}, {'shift': 'Day', 'zone': 'EPIC', 'role': 'Consultant (D1)', 'time': '0800-1730'}, {'shift': 'Day', 'zone': 'Purple', 'role': 'Consultant (D2)', 'time': '0800-1730'}, {'shift': 'Day', 'zone': 'Purple', 'role': 'Consultant (D6)', 'time': '0800-1730'}, {'shift': 'Day', 'zone': 'Purple', 'role': 'Consultant (Dx)', 'time': '0800-1730'}, {'shift': 'Day', 'zone': 'Purple', 'role': 'Registrar (D1)', 'time': '0800-1730'}, {'shift': 'Day', 'zone': 'Purple', 'role': 'Registrar (D5)', 'time': '0800-1730'}, {'shift': 'Day', 'zone': 'Purple', 'role': 'JMO (DA1)', 'time': '0800-1730'}, {'shift': 'Day', 'zone': 'Purple', 'role': 'JMO (DA4)', 'time': '0800-1730'}, {'shift': 'Day', 'zone': 'Green', 'role': 'Consultant (D3)', 'time': '0800-1730'}, {'shift': 'Day', 'zone': 'Green', 'role': 'Consultant (D7)', 'time': '0800-1730'}, {'shift': 'Day', 'zone': 'Green', 'role': 'Consultant (D9)', 'time': '0800-1730'}, {'shift': 'Day', 'zone': 'Green', 'role': 'Registrar (D2)', 'time': '0800-1730'}, {'shift': 'Day', 'zone': 'Green', 'role': 'JMO (DA2)', 'time': '0800-1730'}, {'shift': 'Day', 'zone': 'Green', 'role': 'JMO (DA5)', 'time': '0800-1730'}, {'shift': 'Day', 'zone': 'Green', 'role': 'JMO (DA6)', 'time': '0800-1730'}, {'shift': 'Day', 'zone': 'Fast Track', 'role': 'Consultant (D4)', 'time': '0800-1730'}, {'shift': 'Day', 'zone': 'Fast Track', 'role': 'Consultant (Dy)', 'time': '0800-1730'}, {'shift': 'Day', 'zone': 'Fast Track', 'role': 'Registrar (D3)', 'time': '0800-1730'}, {'shift': 'Day', 'zone': 'Fast Track', 'role': 'Nurse Prac', 'time': '0800-1630'}, {'shift': 'Day', 'zone': 'Fast Track', 'role': 'Adv Musc Physio', 'time': '0930-1800'}, {'shift': 'Day', 'zone': 'Fast Track', 'role': 'JMO (DA3)', 'time': '0800-1730'}, {'shift': 'Day', 'zone': 'SSU/HUB', 'role': 'Consultant (D5)', 'time': '0800-1730'}, {'shift': 'Day', 'zone': 'SSU/HUB', 'role': 'Consultant (DZ)', 'time': '0800-1730'}, {'shift': 'Day', 'zone': 'SSU/HUB', 'role': 'OM Registrar (D4)', 'time': '0800-1730'}, {'shift': 'Day', 'zone': 'SSU/HUB', 'role': 'SSU JMO (DO1)', 'time': '0800-1730'}, {'shift': 'Day', 'zone': 'SSU/HUB', 'role': 'SSU JMO (DO2)', 'time': '0800-1730'}, {'shift': 'Day', 'zone': 'SSU/HUB', 'role': 'HUB JMO (DO3)', 'time': '0800-1730'}, {'shift': 'Day', 'zone': 'Rapid Stay', 'role': 'JMO (DO4)', 'time': '0800-1730'}, {'shift': 'Swing', 'zone': 'Purple', 'role': 'Consultant (S6)', 'time': '1100-2030'}, {'shift': 'Swing', 'zone': 'Purple', 'role': 'Registrar (S2)', 'time': '1100-2030'}, {'shift': 'Swing', 'zone': 'Purple', 'role': 'Registrar (S5)', 'time': '1100-2030'}, {'shift': 'Swing', 'zone': 'Purple', 'role': 'Registrar (S6)', 'time': '1100-2030'}, {'shift': 'Swing', 'zone': 'Green', 'role': 'Consultant (S7)', 'time': '1100-2030'}, {'shift': 'Swing', 'zone': 'Green', 'role': 'Registrar (S3)', 'time': '1100-2030'}, {'shift': 'Swing', 'zone': 'Green', 'role': 'Registrar (S4)', 'time': '1100-2030'}, {'shift': 'Swing', 'zone': 'Green', 'role': 'Registrar (S1)', 'time': '1100-2031'}, {'shift': 'Swing', 'zone': 'Fast Track', 'role': 'Adv Musc Physio', 'time': '1030-1730'}, {'shift': 'Swing', 'zone': 'Fast Track', 'role': 'Nurse Prac', 'time': '1000-1830'}, {'shift': 'Eve', 'zone': 'EPIC', 'role': 'Consultant (E1)', 'time': '1530-0100'}, {'shift': 'Eve', 'zone': 'Purple', 'role': 'Consultant (E2)', 'time': '1430-2400'}, {'shift': 'Eve', 'zone': 'Purple', 'role': 'Consultant (E7)', 'time': '1430-2400'}, {'shift': 'Eve', 'zone': 'Purple', 'role': 'Registrar (E1)', 'time': '1430-2400'}, {'shift': 'Eve', 'zone': 'Purple', 'role': 'Registrar (E5)', 'time': '1430-2400'}, {'shift': 'Eve', 'zone': 'Purple', 'role': 'Registrar (E7)', 'time': '1430-2400'}, {'shift': 'Eve', 'zone': 'Purple', 'role': 'Registrar (E9)', 'time': '1430-2400'}, {'shift': 'Eve', 'zone': 'Purple', 'role': 'Registrar (E01)', 'time': '1430-2400'}, {'shift': 'Eve', 'zone': 'Purple', 'role': 'JMO (EA1)', 'time': '1430-2400'}, {'shift': 'Eve', 'zone': 'Purple', 'role': 'JMO (EA4)', 'time': '1430-2200'}, {'shift': 'Eve', 'zone': 'Green', 'role': 'Consultant (E3)', 'time': '1430-2400'}, {'shift': 'Eve', 'zone': 'Green', 'role': 'Consultant (E6)', 'time': '1430-2400'}, {'shift': 'Eve', 'zone': 'Green', 'role': 'Consultant (E9)', 'time': '1430-2400'}, {'shift': 'Eve', 'zone': 'Green', 'role': 'Registrar (E2)', 'time': '1430-2400'}, {'shift': 'Eve', 'zone': 'Green', 'role': 'Registrar (E4)', 'time': '1430-2400'}, {'shift': 'Eve', 'zone': 'Green', 'role': 'Registrar (E8)', 'time': '1430-2400'}, {'shift': 'Eve', 'zone': 'Green', 'role': 'Registrar (Ez)', 'time': '1430-2400'}, {'shift': 'Eve', 'zone': 'Green', 'role': 'Registrar (E02)', 'time': '1430-2400'}, {'shift': 'Eve', 'zone': 'Green', 'role': 'JMO (EA2)', 'time': '1430-2400'}, {'shift': 'Eve', 'zone': 'Green', 'role': 'JMO (EA5)', 'time': '1430-2400'}, {'shift': 'Eve', 'zone': 'Green', 'role': 'JMO (EA6)', 'time': '1430-2400'}, {'shift': 'Eve', 'zone': 'Fast Track', 'role': 'Consultant (E4)', 'time': '1430-2400'}, {'shift': 'Eve', 'zone': 'Fast Track', 'role': 'Registrar (E6)', 'time': '1430-2400'}, {'shift': 'Eve', 'zone': 'Fast Track', 'role': 'Registrar (E3)', 'time': '1430-2400'}, {'shift': 'Eve', 'zone': 'Fast Track', 'role': 'Nurse Prac', 'time': '1400-2230'}, {'shift': 'Eve', 'zone': 'Fast Track', 'role': 'Adv Musc Physio', 'time': '1400-2200'}, {'shift': 'Eve', 'zone': 'Fast Track', 'role': 'JMO (EA3)', 'time': '1430-2400'}, {'shift': 'Eve', 'zone': 'SSU/HUB', 'role': 'Consultant (E5)', 'time': '1430-2400'}, {'shift': 'Eve', 'zone': 'SSU/HUB', 'role': 'SSU JMO (EO1)', 'time': '1430-2400'}, {'shift': 'Eve', 'zone': 'SSU/HUB', 'role': 'SSU JMO (EO2)', 'time': '1430-2400'}, {'shift': 'Eve', 'zone': 'SSU/HUB', 'role': 'HUB JMO (EO3)', 'time': '1430-2400'}, {'shift': 'Eve', 'zone': 'Rapid Stay', 'role': 'JMO (EO4)', 'time': '1430-2400'}, {'shift': 'Night', 'zone': None, 'role': 'FACEM on call', 'time': '0100-0430'}, {'shift': 'Night', 'zone': 'ERIC', 'role': 'Registrar (N1)', 'time': '2300-0830'}, {'shift': 'Night', 'zone': 'Purple', 'role': 'Registrar (N2)', 'time': '2300-0830'}, {'shift': 'Night', 'zone': 'Purple', 'role': 'Registrar (N6)', 'time': '2300-0830'}, {'shift': 'Night', 'zone': 'Purple', 'role': 'JMO (NA1)', 'time': '2300-0830'}, {'shift': 'Night', 'zone': 'Green', 'role': 'Registrar (N3)', 'time': '2300-0830'}, {'shift': 'Night', 'zone': 'Green', 'role': 'JMO (NA3)', 'time': '2300-0830'}, {'shift': 'Night', 'zone': 'Fast Track', 'role': 'Registrar (N5)', 'time': '2300-0830'}, {'shift': 'Night', 'zone': 'Fast Track', 'role': 'JMO (NA2)', 'time': '2300-0830'}, {'shift': 'Night', 'zone': 'SSU/HUB', 'role': 'Registrar (N4)', 'time': '2300-0830'}, {'shift': 'Night', 'zone': 'SSU/HUB', 'role': 'JMO (NO1)', 'time': '2300-0830'}, {'shift': 'Night', 'zone': 'SSU/HUB', 'role': 'JMO (NO2)', 'time': '2300-0830'}, {'shift': 'Night', 'zone': 'Rapid Stay', 'role': 'JMO (NO3)', 'time': '2300-0830'}]


# ============================================================
# STEP 5 - Lookup functions (same logic as get_today_from_source.py)
# ============================================================

VALID_CONSULTANT_CODES = set()
VALID_REGISTRAR_CODES = set()
for _row in TEMPLATE:
    _m = re.match(r"^(.*?)\s*\(([^)]+)\)\s*$", _row["role"])
    if _m:
        _prefix, _code = _m.group(1).strip().lower(), _m.group(2).strip().lower()
        if _prefix == "consultant":
            VALID_CONSULTANT_CODES.add(_code)
        elif _prefix in ("registrar", "om registrar"):
            VALID_REGISTRAR_CODES.add(_code)

NON_ZONE_CODES = {"x", "l", "8", "rt", "cl", "sl", "bl", "el"}

def is_non_zone_code(stripped_lower):
    """Codes that are recognized but never belong in a clinical zone
    (leave, training, admin day, etc.) - these should NOT be flagged
    as 'unallocated' even though they don't match a template slot."""
    return stripped_lower in NON_ZONE_CODES or stripped_lower.startswith("pl")

def code_period(stripped_lower):
    m = re.match(r"^([dsen])", stripped_lower)
    return {"d": "Day", "s": "Swing", "e": "Eve", "n": "Night"}.get(m.group(1)) if m else None

def strip_suffix(code):
    c = code.strip()
    if c.lower().endswith("-t"):
        c = c[:-2]
    if c and c[-1].lower() in ("a", "c", "p") and len(c) > 1:
        c = c[:-1]
    return c

NAME_SUFFIX_RE = re.compile(r"^(.*?)\s*(\([^)]*\))\s*$")

def reverse_registrar_name(n):
    n = n.strip()
    m = NAME_SUFFIX_RE.match(n)
    suffix, core = "", n
    if m:
        core = m.group(1).strip()
        suffix = " " + m.group(2)
    parts = core.split()
    if len(parts) == 2:
        return parts[1] + " " + parts[0] + suffix
    if len(parts) <= 1:
        return core + suffix
    return parts[-1] + " " + " ".join(parts[:-1]) + suffix

ROLE_RE = re.compile(r"^(.*?)\s*\(([^)]+)\)\s*$")

def classify_role(role_label):
    m = ROLE_RE.match(role_label)
    if not m:
        return None, None
    prefix, code = m.group(1).strip(), m.group(2).strip()
    low = prefix.lower()
    if low == "consultant":
        return "consultant", code.lower()
    if low in ("registrar", "om registrar"):
        return "registrar", code
    if "jmo" in low:
        return "jmo", code
    return None, None


from types import SimpleNamespace

def make_lookups(consultant_periods, registrar_periods, jmo_people, np_people, amp_people):
    initials_map = {}
    for p in consultant_periods:
        for c in p["c"]:
            parts = c["n"].split()
            if len(parts) >= 2:
                initials_map.setdefault((parts[0][0] + parts[-1][0]).upper(), c["n"])

    def consultant_period_for(dt):
        for p in consultant_periods:
            if p["c"] and p["d"]:
                dates = p["d"].split("|")
                if dt in dates:
                    return p, dates.index(dt)
        return None, None

    def registrar_period_for(dt):
        for p in registrar_periods:
            dates = p["d"].split("|")
            if dt in dates:
                return p, dates.index(dt)
        return None, None

    def consultant_lookup(dt, code):
        p, idx = consultant_period_for(dt)
        if p is None:
            return ""
        target = code.lower()
        for c in p["c"]:
            codes = c["v"].split("|")
            if idx < len(codes) and strip_suffix(codes[idx]).lower() == target:
                return c["n"]
        return ""

    def registrar_lookup(dt, code):
        p, idx = registrar_period_for(dt)
        if p is None:
            return ""
        target = code.lower()
        for c in p["c"]:
            codes = c["v"].split("|")
            if idx < len(codes) and strip_suffix(codes[idx]).lower() == target:
                return reverse_registrar_name(c["n"])
        return ""

    def jmo_lookup(dt, code):
        target = code.lower()
        for name, codes in jmo_people.items():
            raw = codes.get(dt, "")
            if raw and strip_suffix(raw).lower() == target:
                core = raw.strip().lower()
                if core.endswith("-t"):
                    core = core[:-2]
                has_p = core.endswith("p")
                return ("~" + name) if has_p else name
        return ""

    def npamp_lookup(dt, shift_name, people_dict):
        letter = {"Day": "d", "Swing": "s", "Eve": "e", "Night": "n"}.get(shift_name, "")
        matches = [name for name, codes in people_dict.items()
                   if codes.get(dt, "").strip().lower() == letter]
        return ", ".join(matches)

    def facem_oncall_lookup(dt, which):
        p, idx = consultant_period_for(dt)
        if p is None:
            return ""
        prefix = "d" if which == "day" else "e"
        for c in p["c"]:
            codes = c["v"].split("|")
            if idx >= len(codes):
                continue
            code = codes[idx].lower()
            if code.startswith(prefix) and code.endswith("c") and len(code) > 2:
                return c["n"]
        return ""

    def director_oncall_lookup(dt):
        p, idx = consultant_period_for(dt)
        if p is None or "oncall" not in p["m"]:
            return ""
        vals = p["m"]["oncall"].split("|")
        if idx >= len(vals):
            return ""
        initials = vals[idx].strip()
        return initials_map.get(initials, initials)

    def next_day(dt):
        return (date.fromisoformat(dt) + timedelta(days=1)).isoformat()

    def find_unallocated(dt):
        """Returns {'Consultants': [names], 'Registrars': [names]} for anyone
        with a real shift code that day that doesn't match any known template
        position (and isn't a recognized non-zone status like leave/training)."""
        result = {"Consultants": {}, "Registrars": {}}
        p, idx = consultant_period_for(dt)
        if p is not None:
            for c in p["c"]:
                codes = c["v"].split("|")
                if idx >= len(codes):
                    continue
                raw = codes[idx]
                if not raw:
                    continue
                stripped = strip_suffix(raw).lower()
                if is_non_zone_code(stripped):
                    continue
                if stripped in VALID_CONSULTANT_CODES:
                    continue
                period = code_period(stripped)
                if period:
                    result["Consultants"].setdefault(period, []).append(c["n"])

        p, idx = registrar_period_for(dt)
        if p is not None:
            for c in p["c"]:
                codes = c["v"].split("|")
                if idx >= len(codes):
                    continue
                raw = codes[idx]
                if not raw:
                    continue
                stripped = strip_suffix(raw).lower()
                if is_non_zone_code(stripped):
                    continue
                if stripped in VALID_REGISTRAR_CODES:
                    continue
                period = code_period(stripped)
                if period:
                    result["Registrars"].setdefault(period, []).append(reverse_registrar_name(c["n"]))
        return result

    return SimpleNamespace(
        consultant_lookup=consultant_lookup,
        registrar_lookup=registrar_lookup,
        jmo_lookup=jmo_lookup,
        npamp_lookup=npamp_lookup,
        facem_oncall_lookup=facem_oncall_lookup,
        director_oncall_lookup=director_oncall_lookup,
        find_unallocated=find_unallocated,
        next_day=next_day,
    )


def build_today_data(consultant_periods, registrar_periods, jmo_people, np_people, amp_people, comments=None):
    comments = comments or {}
    L = make_lookups(consultant_periods, registrar_periods, jmo_people, np_people, amp_people)
    consultant_lookup = L.consultant_lookup
    registrar_lookup = L.registrar_lookup
    jmo_lookup = L.jmo_lookup
    npamp_lookup = L.npamp_lookup
    facem_oncall_lookup = L.facem_oncall_lookup
    director_oncall_lookup = L.director_oncall_lookup
    next_day = L.next_day

    def build_day(dt):
        oncall = director_oncall_lookup(dt)
        shifts = []
        shift_index = {}
        pending_day_facem_note = None

        for meta in TEMPLATE:
            shift_name = meta["shift"]
            if shift_name not in shift_index:
                shift_index[shift_name] = {"name": shift_name, "top_notes": [], "zones": [], "zone_index": {}}
                shifts.append(shift_index[shift_name])
            entry = shift_index[shift_name]

            if meta["role"] == "FACEM on call" and meta["zone"] is None:
                which = "day" if meta["time"] == "0430-0800" else "night"
                if which == "night":
                    name = facem_oncall_lookup(dt, which)
                    if name:
                        entry["top_notes"].append(["FACEM on Call, " + meta["time"], name])
                else:
                    name = facem_oncall_lookup(dt, which)
                    if name:
                        pending_day_facem_note = ["FACEM on Call, " + meta["time"], name]
                continue

            role_label = meta["role"]
            role_type, code = classify_role(role_label)

            if role_label == "Nurse Prac":
                value, generic_role = npamp_lookup(dt, shift_name, np_people), "Nurse Prac"
            elif role_label == "Adv Musc Physio":
                value, generic_role = npamp_lookup(dt, shift_name, amp_people), "Adv Musc Physio"
            elif role_type == "consultant":
                value, generic_role = consultant_lookup(dt, code), "Consultants"
            elif role_type == "registrar":
                value, generic_role = registrar_lookup(dt, code), "Registrars"
            elif role_type == "jmo":
                value, generic_role = jmo_lookup(dt, code), "JMO"
            else:
                value, generic_role = "", role_label

            if not value:
                continue

            zone_name = meta["zone"] or "Other"
            if zone_name not in entry["zone_index"]:
                entry["zone_index"][zone_name] = {"zone": zone_name, "roles": [], "role_lookup": {}}
                entry["zones"].append(entry["zone_index"][zone_name])
            zone = entry["zone_index"][zone_name]

            if generic_role not in zone["role_lookup"]:
                role_entry = {"role": generic_role, "names": []}
                zone["role_lookup"][generic_role] = role_entry
                zone["roles"].append(role_entry)
            for nm in [x.strip() for x in value.split(",") if x.strip()]:
                if nm not in zone["role_lookup"][generic_role]["names"]:
                    zone["role_lookup"][generic_role]["names"].append(nm)

        unallocated = L.find_unallocated(dt)
        for role_key, generic_role in (("Consultants", "Consultants"), ("Registrars", "Registrars")):
            for shift_name, names in unallocated[role_key].items():
                if shift_name not in shift_index:
                    shift_index[shift_name] = {"name": shift_name, "top_notes": [], "zones": [], "zone_index": {}}
                    shifts.append(shift_index[shift_name])
                entry = shift_index[shift_name]
                if "Unallocated" not in entry["zone_index"]:
                    entry["zone_index"]["Unallocated"] = {"zone": "Unallocated", "roles": [], "role_lookup": {}}
                    entry["zones"].append(entry["zone_index"]["Unallocated"])
                zone = entry["zone_index"]["Unallocated"]
                if generic_role not in zone["role_lookup"]:
                    role_entry = {"role": generic_role, "names": []}
                    zone["role_lookup"][generic_role] = role_entry
                    zone["roles"].append(role_entry)
                for nm in names:
                    if nm not in zone["role_lookup"][generic_role]["names"]:
                        zone["role_lookup"][generic_role]["names"].append(nm)

        if pending_day_facem_note and "Night" in shift_index:
            shift_index["Night"]["top_notes"].append(pending_day_facem_note)

        for s in shifts:
            del s["zone_index"]
            for z in s["zones"]:
                del z["role_lookup"]
            remaining_zones = []
            for z in s["zones"]:
                if z["zone"].strip().lower() in ("epic", "eric"):
                    names = [n for role in z["roles"] for n in role["names"]]
                    if names:
                        s["top_notes"].append(["EPIC", ", ".join(names)])
                    continue
                remaining_zones.append(z)
            s["zones"] = [z for z in remaining_zones if z["roles"]]

        date_obj = date.fromisoformat(dt)
        return {"date": date_obj.strftime("%A %d %B"), "oncall": oncall,
                "comment": comments.get(dt, ""), "shifts": shifts}

    return {"panel1": build_day(TODAY), "panel2": build_day(TOMORROW)}


def build_my_roster_data(consultant_periods, registrar_periods):
    consultants = {}
    for p in consultant_periods:
        dates = p["d"].split("|")
        for c in p["c"]:
            codes = c["v"].split("|")
            bucket = consultants.setdefault(c["n"], {})
            for d, code in zip(dates, codes):
                if code and d:
                    bucket[d] = code

    registrars = {}
    for p in registrar_periods:
        dates = p["d"].split("|")
        for c in p["c"]:
            codes = c["v"].split("|")
            bucket = registrars.setdefault(c["n"], {})
            for d, code in zip(dates, codes):
                if code and d:
                    bucket[d] = code

    window_start = (local_today() - timedelta(days=30)).isoformat()

    def pack(name_dict, role):
        out = []
        for name, shifts in sorted(name_dict.items()):
            items = sorted((d, c) for d, c in shifts.items() if d >= window_start)
            if not items:
                continue
            has_real_shift = any(c.strip().lower() != "x" for _, c in items)
            if not has_real_shift:
                continue
            out.append({
                "n": name, "r": role,
                "d": "|".join(d for d, _ in items),
                "v": "|".join(c for _, c in items),
            })
        return out

    return pack(consultants, "C") + pack(registrars, "R")


def build_weekly_data(consultant_periods, registrar_periods, jmo_people, np_people, amp_people, n_weeks=8):
    L = make_lookups(consultant_periods, registrar_periods, jmo_people, np_people, amp_people)

    def build_week(dates):
        oncall = [L.director_oncall_lookup(d) for d in dates]

        shifts = []
        shift_index = {}
        for meta in TEMPLATE:
            shift_name = meta["shift"]
            if shift_name not in shift_index:
                shift_index[shift_name] = {"name": shift_name, "notes": [], "zones": [], "zone_index": {}}
                shifts.append(shift_index[shift_name])
            entry = shift_index[shift_name]

            if meta["role"] == "FACEM on call" and meta["zone"] is None:
                which = "day" if meta["time"] == "0430-0800" else "night"
                values = [L.facem_oncall_lookup(d, which) for d in dates]
                if any(values):
                    entry["notes"].append({"label": "FACEM on Call, " + meta["time"], "values": values})
                continue

            role_label = meta["role"]
            role_type, code = classify_role(role_label)

            if role_label == "Nurse Prac":
                values = [L.npamp_lookup(d, shift_name, np_people) for d in dates]
            elif role_label == "Adv Musc Physio":
                values = [L.npamp_lookup(d, shift_name, amp_people) for d in dates]
            elif role_type == "consultant":
                values = [L.consultant_lookup(d, code) for d in dates]
            elif role_type == "registrar":
                values = [L.registrar_lookup(d, code) for d in dates]
            elif role_type == "jmo":
                values = [L.jmo_lookup(d, code) for d in dates]
            else:
                values = ["" for _ in dates]

            if not any(values):
                continue

            zone_name = meta["zone"] or "Other"
            if zone_name not in entry["zone_index"]:
                entry["zone_index"][zone_name] = {"zone": zone_name, "roles": []}
                entry["zones"].append(entry["zone_index"][zone_name])
            entry["zone_index"][zone_name]["roles"].append({"role": role_label, "values": values})

        for s in shifts:
            del s["zone_index"]

        unallocated_per_day = [L.find_unallocated(d) for d in dates]
        for role_key, generic_role in (("Consultants", "Consultants"), ("Registrars", "Registrars")):
            all_shift_names = set()
            for day_result in unallocated_per_day:
                all_shift_names.update(day_result[role_key].keys())
            for shift_name in all_shift_names:
                values = [", ".join(day_result[role_key].get(shift_name, [])) for day_result in unallocated_per_day]
                if not any(values):
                    continue
                if shift_name not in shift_index:
                    shift_index[shift_name] = {"name": shift_name, "notes": [], "zones": [], "zone_index": {}}
                    shifts.append(shift_index[shift_name])
                entry = shift_index[shift_name]
                unalloc_zone = next((z for z in entry["zones"] if z["zone"] == "Unallocated"), None)
                if unalloc_zone is None:
                    unalloc_zone = {"zone": "Unallocated", "roles": []}
                    entry["zones"].append(unalloc_zone)
                unalloc_zone["roles"].append({"role": generic_role, "values": values})

        for s in shifts:
            s.pop("zone_index", None)

        return {"d": "|".join(dates), "oncall": "|".join(oncall), "shifts": shifts}

    today = local_today()
    monday_start = today - timedelta(days=today.weekday())

    weeks = []
    for i in range(n_weeks):
        week_start = monday_start + timedelta(days=7 * i)
        week_dates = [(week_start + timedelta(days=j)).isoformat() for j in range(7)]
        weeks.append(build_week(week_dates))
    return weeks


def extract_amp2026(wb):
    ws = wb["AMP 2026"]

    def extract_block(title_row, date_row, name_row_start, name_row_end, col_start, col_end):
        title = fmt_code(ws.cell(row=title_row, column=1).value)
        m = re.search(r"(20\d\d)", title)
        target_year = int(m.group(1)) if m else None

        raw_dates = []
        for c in range(col_start, col_end + 1):
            v = ws.cell(row=date_row, column=c).value
            if isinstance(v, datetime):
                raw_dates.append((c, v))
        if not raw_dates or target_year is None:
            return {}

        # self-correcting: figure out the offset from whatever year is actually
        # in the sheet right now, rather than hard-coding a fixed correction -
        # so this keeps working even if the stale dates get fixed later.
        raw_year = raw_dates[0][1].year
        offset = target_year - raw_year

        people = {}
        for r in range(name_row_start, name_row_end + 1):
            name = fmt_code(ws.cell(row=r, column=1).value)
            if not name:
                continue
            codes = {}
            for c, dv in raw_dates:
                corrected = dv.replace(year=dv.year + offset)
                code = fmt_code(ws.cell(row=r, column=c).value)
                if code and code != "0":
                    codes[corrected.strftime("%Y-%m-%d")] = code
            people.setdefault(name, {}).update(codes)
        return people

    block1 = extract_block(1, 2, 3, 13, 2, 182)
    block2 = extract_block(15, 16, 17, 27, 2, 185)

    amp2026 = {}
    for name, codes in block1.items():
        amp2026.setdefault(name, {}).update(codes)
    for name, codes in block2.items():
        amp2026.setdefault(name, {}).update(codes)
    return amp2026


def build_amp_roster_data(amp2026_people):
    all_dates = sorted({d for codes in amp2026_people.values() for d in codes})
    if not all_dates:
        return []
    year = int(all_dates[0][:4])

    months = []
    for m in range(1, 13):
        _, days_in_month = __import__("calendar").monthrange(year, m)
        month_dates = [date(year, m, d).isoformat() for d in range(1, days_in_month + 1)]
        people = []
        for name in sorted(amp2026_people.keys()):
            codes = amp2026_people[name]
            values = [codes.get(d, "") for d in month_dates]
            if any(values):
                people.append({"n": name, "v": "|".join(values)})
        months.append({
            "label": date(year, m, 1).strftime("%B %Y"),
            "d": "|".join(month_dates),
            "c": people,
        })
    return months


def extract_comments(wb):
    if "Comments" not in wb.sheetnames:
        return {}
    ws = wb["Comments"]
    comments = {}
    for r in range(2, ws.max_row + 1):
        dv = ws.cell(row=r, column=1).value
        comment = fmt_code(ws.cell(row=r, column=3).value)
        if isinstance(dv, datetime) and comment:
            comments[dv.strftime("%Y-%m-%d")] = comment
    return comments


# ============================================================
# MAIN
# ============================================================

def main():
    print("Fetching access token...")
    access_token = get_access_token()

    print("Downloading roster file...")
    file_bytes = download_roster(access_token)

    print("Loading workbook...")
    wb = openpyxl.load_workbook(file_bytes, data_only=True)

    print("Extracting consultant periods...")
    consultant_periods = extract_all_consultants(wb)
    print(f"  {len(consultant_periods)} periods found")

    print("Extracting registrar terms...")
    registrar_periods = extract_all_registrars(wb)
    print(f"  {len(registrar_periods)} terms found")

    print("Extracting JMO/NP/AMP...")
    jmo_people, np_people, amp_people = extract_jmo_np(wb)
    print(f"  {len(jmo_people)} JMOs, {len(np_people)} NPs, {len(amp_people)} AMPs")

    print("Extracting Comments tab...")
    comments = extract_comments(wb)
    print(f"  {len(comments)} dated comments found")

    print("Building today/tomorrow view...")
    data = build_today_data(consultant_periods, registrar_periods, jmo_people, np_people, amp_people, comments)

    print("Building My Roster (per-person) view...")
    my_roster_data = build_my_roster_data(consultant_periods, registrar_periods)
    print(f"  {len(my_roster_data)} people with recent/upcoming shifts")

    print("Building Weekly Allocations view...")
    weekly_data = build_weekly_data(consultant_periods, registrar_periods, jmo_people, np_people, amp_people)
    print(f"  {len(weekly_data)} weeks")

    print("Extracting AMP 2026 tab...")
    amp2026_people = extract_amp2026(wb)
    amp_roster_data = build_amp_roster_data(amp2026_people)
    print(f"  {len(amp2026_people)} AMPs")

    with open("roster_data.json", "w") as f:
        json.dump(data, f, indent=2)

    with open("consultant_roster.json", "w") as f:
        json.dump(consultant_periods, f, separators=(",", ":"))

    with open("registrar_roster.json", "w") as f:
        json.dump(registrar_periods, f, separators=(",", ":"))

    with open("my_roster.json", "w") as f:
        json.dump(my_roster_data, f, separators=(",", ":"))

    with open("weekly_allocations.json", "w") as f:
        json.dump(weekly_data, f, separators=(",", ":"))

    with open("amp_roster.json", "w") as f:
        json.dump(amp_roster_data, f, separators=(",", ":"))

    print("Done. Wrote all 6 data files.")


if __name__ == "__main__":
    main()
